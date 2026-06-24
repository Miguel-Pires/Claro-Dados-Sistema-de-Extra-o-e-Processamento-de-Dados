"""
Gera Excel no layout exato do modelo DETALHAMENTO (02213268000137).xlsx.

Colunas preenchidas pelo extrator:
  A  Número    — "(11) 91338 6865"
  B  Consumo   — "Sim"
  C  Nome      — "Plugin Smartphone" | "Tablet e Modem"
  D  Valor     — 21.37
  E  Total GB  — "900GB"
  N  Total     — 21.37

Colunas F-M ficam em branco (dados não disponíveis na fatura resumo).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.line import PhoneLine

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Cores exatas do modelo (sem o prefixo FF do alpha)
# ---------------------------------------------------------------------------
_F_MED_BLUE  = PatternFill("solid", fgColor="FF729FCF")  # A1:B1, E1, M1, N1
_F_LITE_BLUE = PatternFill("solid", fgColor="FFB4C7DC")  # C1:D1 (PLANO)
_F_DARK_BLUE = PatternFill("solid", fgColor="FF3465A4")  # F1:G1 (APARELHO)
_F_VLIT_BLUE = PatternFill("solid", fgColor="FFDEE6EF")  # H1:I1 (DADOS)
_F_GRAY      = PatternFill("solid", fgColor="FFCCCCCC")  # J1:L1 (VOZ)

_FONT_WHITE = Font(bold=True, color="FFFFFF", size=10)
_FONT_DARK  = Font(bold=True, color="000000", size=10)
_FONT_DATA  = Font(size=10)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# (fill, font) para cada coluna do sub-header
_COL_STYLE: dict[str, tuple] = {
    "A": (_F_MED_BLUE,  _FONT_WHITE),
    "B": (_F_MED_BLUE,  _FONT_WHITE),
    "C": (_F_LITE_BLUE, _FONT_DARK),
    "D": (_F_LITE_BLUE, _FONT_DARK),
    "E": (_F_MED_BLUE,  _FONT_WHITE),
    "F": (_F_DARK_BLUE, _FONT_WHITE),
    "G": (_F_DARK_BLUE, _FONT_WHITE),
    "H": (_F_VLIT_BLUE, _FONT_DARK),
    "I": (_F_VLIT_BLUE, _FONT_DARK),
    "J": (_F_GRAY,      _FONT_DARK),
    "K": (_F_GRAY,      _FONT_DARK),
    "L": (_F_GRAY,      _FONT_DARK),
    "M": (_F_MED_BLUE,  _FONT_WHITE),
    "N": (_F_MED_BLUE,  _FONT_WHITE),
}


def _write_headers(ws) -> None:
    # --- Linha 1: grupos (merged) ---
    # A1:B1 — sem texto, azul médio
    ws.merge_cells("A1:B1")
    ws["A1"].fill = _F_MED_BLUE

    # C1:D1 — PLANO, azul claro
    ws.merge_cells("C1:D1")
    c = ws["C1"]
    c.value = "PLANO"
    c.fill = _F_LITE_BLUE
    c.font = _FONT_DARK
    c.alignment = _CENTER

    # E1 — sozinha, azul médio
    ws["E1"].fill = _F_MED_BLUE

    # F1:G1 — APARELHO, azul escuro
    ws.merge_cells("F1:G1")
    c = ws["F1"]
    c.value = "APARELHO"
    c.fill = _F_DARK_BLUE
    c.font = _FONT_WHITE
    c.alignment = _CENTER

    # H1:I1 — DADOS, azul muito claro
    ws.merge_cells("H1:I1")
    c = ws["H1"]
    c.value = "DADOS"
    c.fill = _F_VLIT_BLUE
    c.font = _FONT_DARK
    c.alignment = _CENTER

    # J1:L1 — VOZ, cinza
    ws.merge_cells("J1:L1")
    c = ws["J1"]
    c.value = "VOZ"
    c.fill = _F_GRAY
    c.font = _FONT_DARK
    c.alignment = _CENTER

    # M1 e N1 — azul médio
    ws["M1"].fill = _F_MED_BLUE
    ws["N1"].fill = _F_MED_BLUE

    # --- Linha 2: sub-cabeçalhos ---
    sub = [
        ("A2", "Número"),
        ("B2", "Consumo"),
        ("C2", "Nome"),
        ("D2", "Valor"),
        ("E2", "Total GB"),
        ("F2", "Nome"),
        ("G2", "Valor"),
        ("H2", "Local"),
        ("I2", "Inter"),
        ("J2", "Local"),
        ("K2", "LD"),
        ("L2", "Inter"),
        ("M2", "Outros"),
        ("N2", "Total"),
    ]
    for coord, label in sub:
        col = coord[0]
        fill, font = _COL_STYLE[col]
        c = ws[coord]
        c.value = label
        c.fill = fill
        c.font = font
        c.alignment = _CENTER


def _set_column_widths(ws) -> None:
    widths = {
        "A": 14.7,
        "B": 15.1,
        "C": 49.1,
        "D": 11.0,
        "E": 12.6,
        "F": 44.1,
        "G": 12.1,
        "H": 12.1,
        "I": 10.0,
        "J": 10.0,
        "K": 10.0,
        "L": 10.0,
        "M": 10.0,
        "N": 10.0,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def export_excel(lines: List[PhoneLine], output_path: str | Path) -> Path:
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Claro Empresas"

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16
    ws.freeze_panes = "A3"

    _write_headers(ws)
    _set_column_widths(ws)

    for idx, line in enumerate(lines, start=3):
        gb_str = line.gb  # já vem como "900GB", "1TB", "15MB", etc.

        ws[f"A{idx}"].value = line.telefone_mask
        ws[f"B{idx}"].value = line.consumo
        ws[f"C{idx}"].value = line.plano
        ws[f"D{idx}"].value = line.valor
        ws[f"E{idx}"].value = gb_str
        ws[f"H{idx}"].value = line.mb_usage  # 0.0 quando sem consumo
        ws[f"H{idx}"].number_format = "#,##0.00"
        ws[f"N{idx}"].value = line.valor

        for col in "ABCDEFGHIJKLMN":
            cell = ws[f"{col}{idx}"]
            cell.font = _FONT_DATA
            if col == "A":
                cell.alignment = _LEFT
            elif col == "C":
                cell.alignment = _LEFT
            else:
                cell.alignment = _CENTER
            if col in ("D", "N") and isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

    wb.save(output_path)
    logger.info("Excel gerado: %s (%d linhas)", output_path.name, len(lines))
    return output_path
